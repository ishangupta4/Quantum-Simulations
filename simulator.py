from qiskit.tools.monitor import job_monitor
from qiskit import QuantumCircuit, execute, Aer, QuantumRegister
from qiskit.tools.jupyter import *
from qiskit.visualization import *
import numpy as np
from qiskit.converters import circuit_to_gate
from qiskit.circuit.library.standard_gates import ZGate
import random
from qiskit import IBMQ, transpile
from openpyxl import Workbook, load_workbook


def oracle(n, targets):
    '''
    Given the number of qubits "n" and a list of target states "targets", the function returns the quantum circuit 
    for the oracle.
    '''
    qbits = QuantumRegister(n, 'q')
    circ = QuantumCircuit(qbits, name="Oracle")

    # Creating a sequential circuit for each target state.
    for target in targets:

        # Each string representing a target state is reversed so that the indexing matches with the indices of
        # the corresponding qubits.
        target = list(target)
        target.reverse()
        target = ''.join(target)

        # Flip zero bits
        for i in range(len(target)):
            if target[i] == '0':
                circ.x(i)

        # Apply a n-1 CZ
        CZ = ZGate().control(n-1)
        circ.append(CZ, qbits)

        # Flip back
        for i in range(len(target)):
            if target[i] == '0':
                circ.x(i)

# Uncomment the following line if you the circuit to be converted to a gate. This step is not essential to the
# function of the quantum circuit.
#    circ = circuit_to_gate(circ)

    return circ


def diffusion(n):
    '''
    Given the number of qubits "n", the function returns the circuit for the n-qubit amplification operation.
    '''
    qbits = QuantumRegister(n, 'q')
    circ = QuantumCircuit(qbits, name="diffusion")

    circ.h(qbits)
    # Call the oracle function with target state set to the n-qubit zero state.
    circ.append(oracle(n, ['0'*n]), qbits)
    circ.h(qbits)

    return circ


def grover(n, P, oracle):
    '''
    The "grover" function returns the quantum circuit to implement Grover's search algorithm for an "n"-qubit 
    system to search for "m" targets. The argument "oracle" is the quantum circuit for the Oracle for the
    target state/states. The argument "threshold" is the user-given threshold on the success probability. The 
    default value of "threshold is zero."
    '''
    qbits = QuantumRegister(n, 'q')
    circ = QuantumCircuit(n, name="Grover")

    # Initial state: Prepare equal superposition
    circ.h(qbits)

    # Append the oracle followed by diffusion operation k number of times.
    for i in range(int(3/P)):
        if random.uniform(0, 1) <= (1-P):
            circ.append(oracle, qbits)
            circ.append(diffusion(n), qbits)
        else:
            break

    return circ


def rand_key(p):
    key1 = ""
    for i in range(p):

        temp = str(random.randint(0, 1))
        key1 += temp
    return(key1)


IBMQ.save_account('API_TOKEN')
provider = IBMQ.load_account()
backend_service = provider.service('backend')
backends = provider.backends()
simulator_backend = provider.get_backend(
    'simulator_mps')  # name of the backend


def get_probability(n, P):
    tar1 = rand_key(n)
    tars = [tar1]
    form = f'{{0:0{str(n)}b}}'
    # A list of all possible orthogonal states
    all_states = [form.format(i) for i in range(2**n)]

    # The circuit for Oracle given the target states in the list tars.
    orc = oracle(n, tars)

    # Construction of the circuit for Grover's search.
    circuit = grover(n, P, orc)

    # Applying projective measurement to each qubit at the end of the circuit.
    circuit.measure_all()

    # Running the "circuit" using the QASM simulator.
    transpiled_grover_circuit = transpile(
        circuit, simulator_backend, optimization_level=3)
    job = simulator_backend.run(transpiled_grover_circuit, shots=1000)
    job_monitor(job, interval=2)
    result = job.result()
    # "count" is a dict of counts after the projective measurements.
    count = result.get_counts(circuit)

    # if tars[0] is present in count then print is value else print 0
    if tars[0] in count.keys():
        return count[tars[0]]/1000
    else:
        return 0


wb = load_workbook("hash_sim.xlsx")  # Name of the excel workbook
ws = wb["Sheet1"]  # worksheet to read
wr = wb["Sheet2"]  # worksheet to write
NArray = []
PArray = []
indArray = []
# for i in range(2, 1530, 100):

# Column A is for N (number of qubits)
# Column B is for P (probability)
for i in range(2, 4):
    indArray.append(i)
    NArray.append(ws["A"+str(i)].value)
    PArray.append(ws["B"+str(i)].value)

# prob will store the result for each simulation
prob = []
for i in range(len(NArray)):
    prob.append(get_probability(NArray[i], PArray[i]))

# writing to the workbook
for i in range(len(indArray)):
    wr["A"+str(indArray[i])].value = prob[i]

wb.save("hash_sim.xlsx")
print(prob)
